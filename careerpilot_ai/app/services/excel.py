from collections import Counter
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from careerpilot_ai.app.schemas import TrackerRecord


class ExcelService:
    headers = [
        "ID", "Date Found", "Date Applied", "Company", "Role", "Portal", "Job Link",
        "Location", "Remote Policy", "Salary Range", "Required Experience", "Role Category",
        "Overall Score", "Skill Score", "Project Score", "Experience Score", "Salary Score",
        "Location Score", "Growth Score", "Recommendation", "Status", "Strong Matches",
        "Missing Skills", "Best Projects", "CV Version", "Application Answer", "Recruiter Message",
        "Follow-up Message", "Follow-up Date", "Notes", "Last Updated",
    ]

    def export(self, records: list[TrackerRecord]) -> bytes:
        workbook = Workbook()
        applications = workbook.active
        applications.title = "Applications"
        self._style_header(applications, self.headers)
        for record in records:
            applications.append([
                record.id, record.date_found, record.date_applied, record.company, record.role,
                record.portal, record.job_link, record.location, record.remote_policy,
                record.salary_range, record.required_experience, record.role_category, record.match_score,
                record.skill_score, record.project_score, record.experience_score, record.salary_score,
                record.location_score, record.growth_score, record.recommendation.value,
                record.status.value, " | ".join(record.strong_matches), ", ".join(record.missing_skills),
                " | ".join(record.best_projects), record.cv_version_name, record.application_answer,
                record.recruiter_message, record.follow_up_message, record.follow_up_date, record.notes,
                record.updated_at.replace(tzinfo=None),
            ])
        applications.freeze_panes = "A2"
        applications.auto_filter.ref = applications.dimensions

        companies = workbook.create_sheet("Companies Found")
        self._style_header(companies, ["Company", "Applications", "Roles"])
        for company in sorted({record.company for record in records}):
            company_records = [record for record in records if record.company == company]
            companies.append([company, len(company_records), ", ".join(sorted({r.role for r in company_records}))])

        versions = workbook.create_sheet("CV Versions")
        self._style_header(versions, ["Version", "Company", "Role", "Status"])
        versions.append(["Base CV (MVP)", "", "", "Source profile"])
        for record in records:
            if record.cv_version_name:
                versions.append([record.cv_version_name, record.company, record.role, record.status.value])

        followups = workbook.create_sheet("Follow-ups")
        self._style_header(followups, ["Application ID", "Company", "Role", "Follow-up Date", "Status"])
        for record in records:
            if record.follow_up_date:
                followups.append([record.id, record.company, record.role, record.follow_up_date, record.status.value])

        analytics = workbook.create_sheet("Analytics")
        self._style_header(analytics, ["Metric", "Value"])
        statuses = Counter(record.status.value for record in records)
        applied = sum(1 for record in records if record.date_applied)
        interviews = sum(count for status, count in statuses.items() if status.startswith("Interview"))
        metrics = [
            ("Total jobs found", len(records)),
            ("Total applied", applied),
            ("Total skipped", statuses["Skipped"]),
            ("Total interviews", interviews),
            ("Total offers", statuses["Offer"]),
            ("Total rejections", statuses["Rejected"]),
            ("Response rate", applied and f"{(interviews + statuses['Offer'] + statuses['Rejected']) / applied:.1%}" or "0.0%"),
            ("Average match score", round(sum(r.match_score for r in records) / len(records), 1) if records else 0),
            ("Best portals", self._top_values(record.portal for record in records)),
            ("Best role categories", self._top_values(record.role_category for record in records)),
            ("Most common missing skills", self._top_values(skill for record in records for skill in record.missing_skills)),
        ]
        for metric in metrics:
            analytics.append(metric)

        for sheet in workbook.worksheets:
            for column in sheet.columns:
                max_length = min(60, max(len(str(cell.value or "")) for cell in column) + 2)
                sheet.column_dimensions[column[0].column_letter].width = max_length

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    @staticmethod
    def _style_header(sheet, headers: list[str]) -> None:
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")

    @staticmethod
    def _top_values(values) -> str:
        counts = Counter(value for value in values if value)
        return ", ".join(f"{name} ({count})" for name, count in counts.most_common(5)) or "No data"
