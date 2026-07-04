from datetime import date

from openpyxl import load_workbook
import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from careerpilot_ai.app.db import Base
from careerpilot_ai.app.schemas import Recommendation, TrackerCreate, TrackerStatus, TrackerUpdate
from careerpilot_ai.app.services.excel import ExcelService
from careerpilot_ai.app.services.tracker import DuplicateApplicationError, TrackerService


def test_tracker_and_excel_round_trip(tmp_path) -> None:
    engine = create_engine(f"sqlite:///{tmp_path / 'test.db'}")
    Base.metadata.create_all(engine)
    tracker = TrackerService()
    with Session(engine) as session:
        saved = tracker.add(session, TrackerCreate(
            company="Example AI", role="AI Engineer", match_score=82,
            recommendation=Recommendation.APPLY, missing_skills=["Docker"],
            role_category="AI Backend", skill_score=80, project_score=90,
            strong_matches=["FastAPI matches the backend requirement."],
            best_projects=["Ghostverse.ai GenAI Storybook Platform"],
            cv_version_name="Shashi_AI_Engineer_Example_AI.docx",
            recruiter_message="Hello recruiter", follow_up_message="Following up",
            application_answer="A reviewed, truthful answer.",
        ))
        assert saved.status == TrackerStatus.READY
        updated = tracker.update(session, saved.id, TrackerUpdate(status=TrackerStatus.APPLIED))
        assert updated is not None
        assert updated.date_applied == date.today()
        records = tracker.list(session)

    content = ExcelService().export(records)
    path = tmp_path / "tracker.xlsx"
    path.write_bytes(content)
    workbook = load_workbook(path)
    assert workbook.sheetnames == [
        "Applications", "Companies Found", "CV Versions", "Follow-ups", "Analytics"
    ]
    assert workbook["Applications"]["D2"].value == "Example AI"
    assert workbook["CV Versions"]["A3"].value == "Shashi_AI_Engineer_Example_AI.docx"
    assert workbook["Analytics"]["A10"].value == "Best portals"


def test_tracker_detects_duplicates_and_filters_due_followups(tmp_path) -> None:
    engine = create_engine(f"sqlite:///{tmp_path / 'duplicates.db'}")
    Base.metadata.create_all(engine)
    tracker = TrackerService()
    with Session(engine) as session:
        saved = tracker.add(session, TrackerCreate(
            company="Example AI",
            role="Backend Engineer",
            job_link="https://example.com/jobs/42/",
            match_score=70,
            recommendation=Recommendation.MAYBE,
            follow_up_date=date.today(),
        ))
        with pytest.raises(DuplicateApplicationError) as duplicate:
            tracker.add(session, TrackerCreate(
                company="  example   ai ",
                role="BACKEND ENGINEER",
                job_link="https://example.com/jobs/42",
                match_score=72,
                recommendation=Recommendation.MAYBE,
            ))
        assert duplicate.value.record_id == saved.id
        assert len(tracker.list(session)) == 1
        assert [record.id for record in tracker.follow_ups_due(tracker.list(session))] == [saved.id]

        tracker.update(session, saved.id, TrackerUpdate(status=TrackerStatus.SKIPPED))
        assert tracker.follow_ups_due(tracker.list(session)) == []
